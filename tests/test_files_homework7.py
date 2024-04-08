import os
import csv
import zipfile
from pypdf import PdfReader
from path import ZIP_FILE
from openpyxl import load_workbook

def test_pdf():
    with zipfile.ZipFile(os.path.join(ZIP_FILE)) as zip_file:
        with zip_file.open('pdf-test.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[2]
            text = page.extract_text()
            assert 'Тестовый PDF файл' in text

def test_csv():
    with zipfile.ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open('Sample-Spreadsheet-100-rows.csv') as csv_file:
            content = csv_file.read().decode('latin-1')  # читаем содержимое файла
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[1]  # получаем вторую строку

            assert second_row[1] == '1.7 Cubic Foot Compact "Cube" Office Refrigerators'
            assert second_row[2] == 'Barry French'



def test_xlsx():
    with zipfile.ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open('file_example_XLSX_10.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active

            assert sheet.cell(row=2, column=2).value == 'Dulce'
            assert sheet.cell(row=2, column=5).value == 'United States'




